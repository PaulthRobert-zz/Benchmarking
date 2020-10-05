#TODO structure this as a module, reference it from main,py

#use this module to get the data from multiple worksheets from an excel file

from openpyxl import load_workbook

#path in project where 2019 data is stored 
filename = 'C:/Users/paul.davis/source/repos/Benchmarking/Data/osr_final_tables2007.xlsx' 

wb = load_workbook(filename)
sheets = wb.sheetnames
#returns the names of the worksheets

for sheet in sheets:
    print(sheet)